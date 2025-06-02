"""FastAPI endpoints for managing custom calculations."""

from fastapi import APIRouter, HTTPException, Depends, Query
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
import json
from datetime import datetime

# Import your models and dependencies
from app.core.database import get_db_session, get_dw_session
from app.models.calculations import (
    SavedCalculation, CalculationTemplate, CalculationRepository,
    CalculationConfigRequest, CalculationConfigResponse, DropdownOption, DropdownData
)
from app.models.dwh_models import Deal, Tranche, TrancheBal
from app.services.calculation_builder import (
    DynamicSubqueryBuilder, CalculationManager, CalculationConfig,
    CalculationType, AggregationLevel
)
from fastapi.responses import StreamingResponse
import io
import csv
import json
from typing import List
from pydantic import BaseModel

router = APIRouter()

# Helper functions
def get_calculation_type_description(calc_type):
    """Get description for calculation type."""
    descriptions = {
        CalculationType.SUM: "Sum all values",
        CalculationType.AVERAGE: "Calculate arithmetic average",
        CalculationType.WEIGHTED_AVERAGE: "Calculate weighted average using specified weight field",
        CalculationType.COUNT: "Count non-null values",
        CalculationType.MIN: "Find minimum value",
        CalculationType.MAX: "Find maximum value",
        CalculationType.RATIO: "Calculate ratio between two fields",
        CalculationType.PERCENTAGE: "Calculate percentage between two fields"
    }
    return descriptions.get(calc_type, "")

def calc_to_response_model(calc: SavedCalculation) -> CalculationConfigResponse:
    """Convert SavedCalculation to response model."""
    return CalculationConfigResponse(
        id=calc.id,
        name=calc.name,
        description=calc.description,
        calculation_type=calc.calculation_type,
        target_field=calc.target_field,
        aggregation_level=calc.aggregation_level,
        weight_field=calc.weight_field,
        denominator_field=calc.denominator_field,
        cycle_filter=calc.cycle_filter,
        filters=calc.filters,
        created_by=calc.created_by,
        created_at=calc.created_at,
        updated_at=calc.updated_at,
        is_public=calc.is_public
    )

def template_to_response_model(template: CalculationTemplate) -> CalculationConfigResponse:
    """Convert CalculationTemplate to response model."""
    return CalculationConfigResponse(
        id=template.id,
        name=template.name,
        description=template.description,
        calculation_type=template.calculation_type,
        target_field=template.target_field,
        aggregation_level=template.aggregation_level,
        weight_field=template.weight_field,
        denominator_field=template.denominator_field,
        cycle_filter=None,  # Templates don't have cycle filters
        filters=template.default_filters,
        created_by="system",
        created_at=datetime.now(),
        updated_at=datetime.now(),
        is_public=True
    )

def request_to_calculation_config(request: CalculationConfigRequest) -> CalculationConfig:
    """Convert API request to CalculationConfig."""
    return CalculationConfig(
        name=request.name,
        calculation_type=CalculationType(request.calculation_type),
        target_field=request.target_field,
        aggregation_level=AggregationLevel(request.aggregation_level),
        weight_field=request.weight_field,
        denominator_field=request.denominator_field,
        filters=request.filters,
        cycle_filter=request.cycle_filter
    )

# Dependency to get current user (implement based on your auth system)
def get_current_user():
    """Get current user from authentication system."""
    # This would integrate with your authentication system
    return "demo_user"

@router.get("/calculations/dropdown-data", response_model=DropdownData)
async def get_dropdown_data(
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """Get all dropdown data needed for the calculation builder UI."""
    
    repo = CalculationRepository(db)
    
    # Get calculation types
    calculation_types = [
        DropdownOption(
            value=calc_type.value,
            label=calc_type.value.replace('_', ' ').title(),
            description=get_calculation_type_description(calc_type)
        )
        for calc_type in CalculationType
    ]
    
    # Get target fields
    target_fields = [
        DropdownOption(value="ending_balance", label="Ending Balance", description="Tranche ending balance amount"),
        DropdownOption(value="principal_release", label="Principal Release", description="Principal release/loss amount"),
        DropdownOption(value="pass_through_rate", label="Pass Through Rate", description="Interest pass-through rate"),
        DropdownOption(value="accrual_days", label="Accrual Days", description="Number of accrual days"),
        DropdownOption(value="interest_distribution", label="Interest Distribution", description="Interest distribution amount"),
        DropdownOption(value="principal_distribution", label="Principal Distribution", description="Principal distribution amount"),
        DropdownOption(value="interest_accrual", label="Interest Accrual", description="Interest accrual amount"),
        DropdownOption(value="interest_shortfall", label="Interest Shortfall", description="Interest shortfall amount"),
    ]
    
    # Get aggregation levels
    aggregation_levels = [
        DropdownOption(value="deal", label="Deal Level", description="Aggregate across all tranches in a deal"),
        DropdownOption(value="tranche", label="Tranche Level", description="Individual tranche calculations"),
    ]
    
    # Get saved calculations for user
    saved_calcs = repo.get_user_calculations(user_id) + repo.get_public_calculations()
    saved_calculations = [calc_to_response_model(calc) for calc in saved_calcs]
    
    # Get templates
    template_records = repo.get_calculation_templates()
    templates = [template_to_response_model(template) for template in template_records]
    
    return DropdownData(
        calculation_types=calculation_types,
        target_fields=target_fields,
        aggregation_levels=aggregation_levels,
        saved_calculations=saved_calculations,
        templates=templates
    )

@router.post("/calculations", response_model=CalculationConfigResponse)
async def create_calculation(
    request: CalculationConfigRequest,
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """Create a new saved calculation."""
    
    try:
        # Validate the calculation configuration
        calc_config = request_to_calculation_config(request)
        
        # Save to database
        repo = CalculationRepository(db)
        saved_calc = repo.save_calculation(
            config=calc_config,
            name=request.name,
            description=request.description,
            created_by=user_id,
            is_public=request.is_public
        )
        
        return calc_to_response_model(saved_calc)
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid calculation configuration: {str(e)}")

@router.get("/calculations", response_model=List[CalculationConfigResponse])
async def get_calculations(
    user_id: str = Depends(get_current_user),
    include_public: bool = Query(True, description="Include public calculations"),
    search: Optional[str] = Query(None, description="Search term"),
    db: Session = Depends(get_db_session)
):
    """Get saved calculations for the current user."""
    
    repo = CalculationRepository(db)
    
    if search:
        calculations = repo.search_calculations(search, user_id if not include_public else None)
    else:
        calculations = repo.get_user_calculations(user_id)
        if include_public:
            calculations += repo.get_public_calculations()
    
    return [calc_to_response_model(calc) for calc in calculations]

@router.get("/calculations/{calc_id}", response_model=CalculationConfigResponse)
async def get_calculation(
    calc_id: int,
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """Get a specific calculation by ID."""
    
    repo = CalculationRepository(db)
    calculation = repo.get_calculation(calc_id)
    
    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")
    
    # Check permissions
    if calculation.created_by != user_id and not calculation.is_public:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return calc_to_response_model(calculation)

@router.put("/calculations/{calc_id}", response_model=CalculationConfigResponse)
async def update_calculation(
    calc_id: int,
    request: CalculationConfigRequest,
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """Update an existing calculation."""
    
    repo = CalculationRepository(db)
    calculation = repo.get_calculation(calc_id)
    
    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")
    
    # Check permissions
    if calculation.created_by != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        # Update the calculation
        updated_calc = repo.update_calculation(
            calc_id,
            name=request.name,
            description=request.description,
            calculation_type=request.calculation_type,
            target_field=request.target_field,
            aggregation_level=request.aggregation_level,
            weight_field=request.weight_field,
            denominator_field=request.denominator_field,
            cycle_filter=request.cycle_filter,
            filters=request.filters,
            is_public=request.is_public
        )
        
        return calc_to_response_model(updated_calc)
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid calculation configuration: {str(e)}")

@router.delete("/calculations/{calc_id}")
async def delete_calculation(
    calc_id: int,
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """Delete a calculation (soft delete)."""
    
    repo = CalculationRepository(db)
    calculation = repo.get_calculation(calc_id)
    
    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")
    
    # Check permissions
    if calculation.created_by != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    repo.delete_calculation(calc_id)
    return {"message": "Calculation deleted successfully"}

@router.post("/calculations/{calc_id}/test")
async def test_calculation(
    calc_id: int,
    cycle_filter: Optional[int] = Query(None, description="Cycle to test with"),
    limit: int = Query(10, description="Limit results for testing"),
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Test a calculation and return sample results."""
    
    repo = CalculationRepository(db)
    calculation = repo.get_calculation(calc_id)
    
    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")
    
    # Check permissions
    if calculation.created_by != user_id and not calculation.is_public:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        # Convert to CalculationConfig
        calc_config = calculation.to_calculation_config()
        
        # Override cycle filter if provided
        if cycle_filter:
            calc_config.cycle_filter = cycle_filter
        
        # Build and execute the calculation using data warehouse session
        manager = CalculationManager(dw_db)
        
        if calc_config.aggregation_level == AggregationLevel.DEAL:
            base_query = manager.create_enhanced_query("deal", [calc_config])
        else:
            base_query = manager.create_enhanced_query("tranche", [calc_config])
        
        # Execute with limit
        results = base_query.limit(limit).all()
        
        # Convert results to serializable format
        result_data = []
        for row in results:
            row_dict = {}
            # Handle SQLAlchemy row objects
            if hasattr(row, '_fields'):
                for column in row._fields:
                    value = getattr(row, column)
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    row_dict[column] = value
            else:
                # Handle simple objects
                for i, desc in enumerate(base_query.column_descriptions):
                    col_name = desc['name']
                    try:
                        value = row[i] if hasattr(row, '__getitem__') else getattr(row, col_name, None)
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        row_dict[col_name] = value
                    except (IndexError, AttributeError):
                        row_dict[col_name] = None
            result_data.append(row_dict)
        
        return {
            "calculation": calc_to_response_model(calculation),
            "sample_results": result_data,
            "result_count": len(result_data),
            "cycle_filter_used": calc_config.cycle_filter
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Calculation test failed: {str(e)}")

@router.post("/calculations/preview-sql")
async def preview_sql(
    request: CalculationConfigRequest,
    user_id: str = Depends(get_current_user),
    dw_db: Session = Depends(get_dw_session)
):
    """Preview the SQL that would be generated for a calculation."""
    
    try:
        # Convert request to config
        calc_config = request_to_calculation_config(request)
        
        # Build the subquery
        builder = DynamicSubqueryBuilder(dw_db)
        subquery = builder.build_calculation_subquery(calc_config)
        
        # Convert to SQL string
        sql_str = str(subquery.compile(compile_kwargs={"literal_binds": True}))
        
        return {
            "sql_preview": sql_str,
            "calculation_name": request.name,
            "calculation_type": request.calculation_type
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to generate SQL preview: {str(e)}")


# Pydantic models for report builder
class ReportRequest(BaseModel):
    calculation_ids: List[int]
    base_aggregation: str  # 'deal' or 'tranche'
    cycle_filter: Optional[int] = None
    report_name: str = "Custom Report"

class ReportPreviewRequest(BaseModel):
    calculation_ids: List[int]
    base_aggregation: str
    cycle_filter: Optional[int] = None

class ReportResult(BaseModel):
    columns: List[str]
    rows: List[List[Any]]
    row_count: int
    report_name: str
    execution_time_ms: int

@router.get("/reports/available-calculations")
async def get_available_calculations(
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """Get all available calculations for building reports."""
    
    try:
        repo = CalculationRepository(db)
        
        # Get user's calculations and public calculations
        user_calcs = repo.get_user_calculations(user_id)
        public_calcs = repo.get_public_calculations()
        
        # Combine and deduplicate
        all_calcs = user_calcs + public_calcs
        seen_ids = set()
        unique_calcs = []
        
        for calc in all_calcs:
            if calc.id not in seen_ids:
                unique_calcs.append(calc)
                seen_ids.add(calc.id)
        
        # Convert to response models
        return [calc_to_response_model(calc) for calc in unique_calcs]
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load calculations: {str(e)}")

@router.post("/reports/preview-sql")
async def preview_report_sql(
    request: ReportPreviewRequest,
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Generate SQL preview for a report without executing it."""
    
    if not request.calculation_ids:
        raise HTTPException(status_code=400, detail="At least one calculation must be selected")
    
    try:
        repo = CalculationRepository(db)
        
        # Get selected calculations
        calculations = []
        for calc_id in request.calculation_ids:
            calc = repo.get_calculation(calc_id)
            if not calc:
                raise HTTPException(status_code=404, detail=f"Calculation {calc_id} not found")
            
            # Check permissions
            if calc.created_by != user_id and not calc.is_public:
                raise HTTPException(status_code=403, detail=f"Access denied to calculation {calc_id}")
            
            calculations.append(calc)
        
        # Validate aggregation levels are compatible
        aggregation_levels = set(calc.aggregation_level for calc in calculations)
        if len(aggregation_levels) > 1:
            levels_str = ", ".join(sorted(aggregation_levels))
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot mix calculations from different aggregation levels in the same report. Found: {levels_str}. Please select calculations from only one aggregation level."
            )
        
        calc_aggregation = list(aggregation_levels)[0]
        if calc_aggregation != request.base_aggregation:
            raise HTTPException(
                status_code=400,
                detail=f"Base aggregation level '{request.base_aggregation}' does not match selected calculations which are '{calc_aggregation}' level. Please adjust the base aggregation level or select different calculations."
            )


        # Convert to CalculationConfig objects
        calc_configs = []
        for calc in calculations:
            config = calc.to_calculation_config()
            # Override cycle filter if specified in request
            if request.cycle_filter:
                config.cycle_filter = request.cycle_filter
            calc_configs.append(config)
        
        # Generate the combined query
        manager = CalculationManager(dw_db)
        
        if request.base_aggregation == "deal":
            enhanced_query = manager.create_enhanced_query("deal", calc_configs)
        else:
            enhanced_query = manager.create_enhanced_query("tranche", calc_configs)
        
        # Convert to SQL string
        sql_str = str(enhanced_query.statement.compile(compile_kwargs={"literal_binds": True}))
        
        return {
            "sql_query": sql_str,
            "calculation_count": len(calc_configs),
            "base_aggregation": request.base_aggregation,
            "cycle_filter": request.cycle_filter
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate SQL preview: {str(e)}")

@router.post("/reports/execute", response_model=ReportResult)
async def execute_report(
    request: ReportRequest,
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Execute a report and return the results."""
    
    if not request.calculation_ids:
        raise HTTPException(status_code=400, detail="At least one calculation must be selected")
    
    try:
        import time
        start_time = time.time()
        
        repo = CalculationRepository(db)
        
        # Get selected calculations
        calculations = []
        for calc_id in request.calculation_ids:
            calc = repo.get_calculation(calc_id)
            if not calc:
                raise HTTPException(status_code=404, detail=f"Calculation {calc_id} not found")
            
            # Check permissions
            if calc.created_by != user_id and not calc.is_public:
                raise HTTPException(status_code=403, detail=f"Access denied to calculation {calc_id}")
            
            calculations.append(calc)
        
        # Validate aggregation levels are compatible
        aggregation_levels = set(calc.aggregation_level for calc in calculations)
        if len(aggregation_levels) > 1:
            raise HTTPException(
                status_code=400, 
                detail="Cannot mix deal-level and tranche-level calculations in the same report"
            )
        
        # Convert to CalculationConfig objects
        calc_configs = []
        for calc in calculations:
            config = calc.to_calculation_config()
            # Override cycle filter if specified in request
            if request.cycle_filter:
                config.cycle_filter = request.cycle_filter
            calc_configs.append(config)
        
        # Generate and execute the combined query
        manager = CalculationManager(dw_db)
        
        if request.base_aggregation == "deal":
            enhanced_query = manager.create_enhanced_query("deal", calc_configs)
        else:
            enhanced_query = manager.create_enhanced_query("tranche", calc_configs)
        
        # Execute the query
        results = enhanced_query.all()
        
        # Extract column names from the query description
        column_descriptions = enhanced_query.column_descriptions
        columns = [desc['name'] for desc in column_descriptions]
        
        # Convert results to rows with proper serialization
        rows = []
        for result in results:
            row = []
            
            # Handle different types of result objects
            if hasattr(result, '_fields'):
                # SQLAlchemy Row object - access by field name
                for col_name in columns:
                    try:
                        value = getattr(result, col_name)
                        # Handle datetime serialization
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        # Handle None values
                        elif value is None:
                            value = None
                        # Convert to basic Python types
                        elif hasattr(value, '__dict__'):
                            # This is a SQLAlchemy model object, extract the primary key or relevant field
                            if hasattr(value, 'dl_nbr'):
                                value = value.dl_nbr
                            elif hasattr(value, 'id'):
                                value = value.id
                            else:
                                value = str(value)
                        row.append(value)
                    except AttributeError:
                        row.append(None)
            else:
                # Handle tuple or list results
                for i, col_desc in enumerate(column_descriptions):
                    try:
                        if hasattr(result, '__getitem__') and i < len(result):
                            value = result[i]
                        else:
                            value = getattr(result, col_desc['name'], None)
                        
                        # Handle datetime serialization
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        # Handle SQLAlchemy objects
                        elif hasattr(value, '__dict__') and not isinstance(value, (str, int, float, bool)):
                            if hasattr(value, 'dl_nbr'):
                                value = value.dl_nbr
                            elif hasattr(value, 'id'):
                                value = value.id
                            else:
                                value = str(value)
                        
                        row.append(value)
                    except (IndexError, AttributeError):
                        row.append(None)
            
            rows.append(row)
        
        execution_time = int((time.time() - start_time) * 1000)
        
        return ReportResult(
            columns=columns,
            rows=rows,
            row_count=len(rows),
            report_name=request.report_name,
            execution_time_ms=execution_time
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to execute report: {str(e)}")

@router.post("/reports/export/csv")
async def export_report_csv(
    request: ReportRequest,
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Export report results as CSV."""
    
    try:
        # Execute the report to get data
        report_result = await execute_report(request, user_id, db, dw_db)
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(report_result.columns)
        
        # Write data rows
        for row in report_result.rows:
            # Convert None values to empty strings for CSV
            csv_row = ['' if cell is None else str(cell) for cell in row]
            writer.writerow(csv_row)
        
        # Create response
        output.seek(0)
        response = StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{request.report_name}.csv"'
            }
        )
        
        return response
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")

@router.post("/reports/export/excel")
async def export_report_excel(
    request: ReportRequest,
    user_id: str = Depends(get_current_user),
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Export report results as Excel."""
    
    try:
        # Execute the report to get data
        report_result = await execute_report(request, user_id, db, dw_db)
        
        # Create Excel file in memory using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return await export_report_csv(request, user_id, db, dw_db)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
        
        for col_idx, header in enumerate(report_result.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data rows
        for row_idx, row_data in enumerate(report_result.rows, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{request.report_name}.xlsx"'
            }
        )
        
        return response
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {str(e)}")

@router.get("/reports/sample-data")
async def get_sample_report_data(
    base_aggregation: str = Query("deal", description="deal or tranche"),
    limit: int = Query(10, description="Number of sample rows"),
    dw_db: Session = Depends(get_dw_session)
):
    """Get sample data to show what reports look like."""
    
    try:
        if base_aggregation == "deal":
            # Sample deal-level data
            query = dw_db.query(
                Deal.dl_nbr,
                Deal.issr_cde,
                Deal.cdi_file_nme
            ).limit(limit)
        else:
            # Sample tranche-level data  
            query = dw_db.query(
                Tranche.dl_nbr,
                Tranche.tr_id,
                Tranche.tr_cusip_id,
                Deal.issr_cde
            ).join(Deal).limit(limit)
        
        results = query.all()
        
        # Convert to response format
        columns = [desc['name'] for desc in query.column_descriptions]
        rows = []
        
        for result in results:
            row = []
            for i in range(len(columns)):
                value = result[i] if hasattr(result, '__getitem__') else getattr(result, columns[i], None)
                if isinstance(value, datetime):
                    value = value.isoformat()
                row.append(value)
            rows.append(row)
        
        return {
            "columns": columns,
            "rows": rows,
            "sample_data": True,
            "base_aggregation": base_aggregation
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get sample data: {str(e)}")