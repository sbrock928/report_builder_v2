"""FastAPI endpoints for report wizard functionality."""

from fastapi import APIRouter, HTTPException, Depends, Query
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
import json
from datetime import datetime
import time

# Import dependencies
from app.core.database import get_db_session, get_dw_session
from app.models.reports import Report, ReportDeal, ReportTranche, ReportField, FilterCondition
from app.models.report_repository import ReportRepository
from app.models.calculations import SavedCalculation, CalculationRepository
from app.models.dwh_models import Deal, Tranche, TrancheBal
from app.models.report_api_models import (
    ReportCreate, ReportUpdate, ReportResponse, ReportSummaryResponse,
    ReportExecuteRequest, ReportExecutionResult, ReportSchemaResponse,
    AvailableField, DealInfo, TrancheInfo, WizardDataResponse,
    FieldSource, ReportScope, FilterOperator, ReportError
)

router = APIRouter()


# === WIZARD DATA ENDPOINTS ===

@router.get("/wizard-data", response_model=WizardDataResponse)
async def get_wizard_data(
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Get all data needed for the wizard UI."""
    
    try:
        # Get available deals
        deals = dw_db.query(Deal).order_by(Deal.dl_nbr).all()
        deal_infos = [
            DealInfo(
                dl_nbr=deal.dl_nbr,
                issr_cde=deal.issr_cde,
                cdi_file_nme=deal.cdi_file_nme,
                CDB_cdi_file_nme=deal.CDB_cdi_file_nme
            )
            for deal in deals
        ]
        
        # Get available fields by scope
        available_fields = _get_available_fields_by_scope(db)
        
        # Get filter operators
        filter_operators = [
            {
                "value": op.value,
                "label": op.value.replace('_', ' ').title(),
                "description": _get_operator_description(op)
            }
            for op in FilterOperator
        ]
        
        return WizardDataResponse(
            available_fields=available_fields,
            deals=deal_infos,
            calculation_types=["sum", "avg", "weighted_avg", "ratio", "percentage"],
            filter_operators=filter_operators
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading wizard data: {str(e)}")


@router.get("/deals", response_model=List[DealInfo])
async def get_deals(
    search: Optional[str] = Query(None, description="Search term for deals"),
    offset: int = Query(0, ge=0, description="Pagination offset"),
    limit: int = Query(100, ge=1, le=1000, description="Pagination limit"),
    dw_db: Session = Depends(get_dw_session)
):
    """Get available deals with optional search and pagination."""
    
    try:
        query = dw_db.query(Deal)
        
        # Apply search filter
        if search:
            search_filter = (
                Deal.dl_nbr.like(f'%{search}%') |
                Deal.issr_cde.ilike(f'%{search}%') |
                Deal.cdi_file_nme.ilike(f'%{search}%')
            )
            query = query.filter(search_filter)
        
        # Apply pagination
        deals = query.order_by(Deal.dl_nbr).offset(offset).limit(limit).all()
        
        return [
            DealInfo(
                dl_nbr=deal.dl_nbr,
                issr_cde=deal.issr_cde,
                cdi_file_nme=deal.cdi_file_nme,
                CDB_cdi_file_nme=deal.CDB_cdi_file_nme
            )
            for deal in deals
        ]
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading deals: {str(e)}")


@router.post("/tranches", response_model=Dict[str, List[TrancheInfo]])
async def get_tranches(
    deal_ids: List[int],
    dw_db: Session = Depends(get_dw_session)
):
    """Get tranches for specified deals."""
    
    try:
        if not deal_ids:
            return {}
        
        tranches = dw_db.query(Tranche).filter(
            Tranche.dl_nbr.in_(deal_ids)
        ).order_by(Tranche.dl_nbr, Tranche.tr_id).all()
        
        # Group tranches by deal
        tranches_by_deal = {}
        for tranche in tranches:
            deal_key = str(tranche.dl_nbr)
            if deal_key not in tranches_by_deal:
                tranches_by_deal[deal_key] = []
            
            tranches_by_deal[deal_key].append(
                TrancheInfo(
                    dl_nbr=tranche.dl_nbr,
                    tr_id=tranche.tr_id,
                    tr_cusip_id=tranche.tr_cusip_id
                )
            )
        
        return tranches_by_deal
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading tranches: {str(e)}")


@router.get("/fields/{scope}", response_model=List[AvailableField])
async def get_available_fields(
    scope: ReportScope,
    db: Session = Depends(get_db_session)
):
    """Get available fields for the specified scope."""
    
    try:
        available_fields = _get_available_fields_by_scope(db)
        return available_fields.get(scope.value, [])
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading fields: {str(e)}")


# === REPORT CRUD ENDPOINTS ===

@router.post("/reports", response_model=ReportResponse)
async def create_report(
    report_data: ReportCreate,
    db: Session = Depends(get_db_session)
):
    """Create a new report configuration."""
    
    try:
        repo = ReportRepository(db)
        
        # Convert Pydantic models to dictionaries
        selected_deals = [deal.dict() for deal in report_data.selected_deals]
        selected_fields = [field.dict() for field in report_data.selected_fields]
        filter_conditions = [condition.dict() for condition in report_data.filter_conditions]
        
        # Create the report
        report = repo.create_report(
            name=report_data.name,
            description=report_data.description,
            scope=report_data.scope.value,
            created_by=report_data.created_by,
            selected_deals=selected_deals,
            selected_fields=selected_fields,
            filter_conditions=filter_conditions
        )
        
        return _convert_report_to_response(report)
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error creating report: {str(e)}")


@router.get("/reports", response_model=List[ReportSummaryResponse])
async def get_reports(
    search: Optional[str] = Query(None, description="Search term"),
    created_by: Optional[str] = Query(None, description="Filter by creator"),
    db: Session = Depends(get_db_session)
):
    """Get all report configurations."""
    
    try:
        repo = ReportRepository(db)
        
        if search:
            summaries = repo.search_reports(search, created_by)
            return [_convert_report_to_summary(report) for report in summaries]
        else:
            summaries = repo.get_report_summaries(created_by)
            return [ReportSummaryResponse(**summary) for summary in summaries]
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading reports: {str(e)}")


@router.get("/reports/{report_id}", response_model=ReportResponse)
async def get_report(
    report_id: int,
    db: Session = Depends(get_db_session)
):
    """Get a specific report configuration."""
    
    try:
        repo = ReportRepository(db)
        report = repo.get_report(report_id)
        
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        
        return _convert_report_to_response(report)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading report: {str(e)}")


@router.put("/reports/{report_id}", response_model=ReportResponse)
async def update_report(
    report_id: int,
    report_data: ReportUpdate,
    db: Session = Depends(get_db_session)
):
    """Update an existing report configuration."""
    
    try:
        repo = ReportRepository(db)
        
        # Convert Pydantic models to dictionaries if provided
        update_data = {}
        if report_data.name is not None:
            update_data['name'] = report_data.name
        if report_data.description is not None:
            update_data['description'] = report_data.description
        if report_data.scope is not None:
            update_data['scope'] = report_data.scope.value
        if report_data.selected_deals is not None:
            update_data['selected_deals'] = [deal.dict() for deal in report_data.selected_deals]
        if report_data.selected_fields is not None:
            update_data['selected_fields'] = [field.dict() for field in report_data.selected_fields]
        if report_data.filter_conditions is not None:
            update_data['filter_conditions'] = [condition.dict() for condition in report_data.filter_conditions]
        
        # Update the report
        report = repo.update_report(report_id, **update_data)
        
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        
        return _convert_report_to_response(report)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error updating report: {str(e)}")


@router.delete("/reports/{report_id}")
async def delete_report(
    report_id: int,
    db: Session = Depends(get_db_session)
):
    """Delete a report configuration."""
    
    try:
        repo = ReportRepository(db)
        success = repo.delete_report(report_id)
        
        if not success:
            raise HTTPException(status_code=404, detail="Report not found")
        
        return {"message": "Report deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting report: {str(e)}")


@router.get("/reports/{report_id}/schema", response_model=ReportSchemaResponse)
async def get_report_schema(
    report_id: int,
    db: Session = Depends(get_db_session)
):
    """Get report schema for preview purposes."""
    
    try:
        repo = ReportRepository(db)
        report = repo.get_report(report_id)
        
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        
        # Generate column definitions
        columns = []
        skeleton_data = []
        
        for field in report.selected_fields:
            columns.append({
                "field": field.field_name,
                "header": field.display_name,
                "type": field.field_type
            })
        
        # Generate skeleton data (3 sample rows)
        for i in range(3):
            row = {}
            for field in report.selected_fields:
                if field.field_type == "number":
                    row[field.field_name] = f"sample_number_{i}"
                elif field.field_type == "percentage":
                    row[field.field_name] = f"sample_%_{i}"
                elif field.field_type == "date":
                    row[field.field_name] = f"sample_date_{i}"
                else:
                    row[field.field_name] = f"sample_text_{i}"
            skeleton_data.append(row)
        
        return ReportSchemaResponse(
            report_id=report.id,
            title=report.name,
            columns=columns,
            skeleton_data=skeleton_data
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating report schema: {str(e)}")


# === REPORT EXECUTION ENDPOINTS ===

@router.post("/reports/{report_id}/execute", response_model=ReportExecutionResult)
async def execute_report(
    report_id: int,
    cycle_filter: Optional[int] = Query(None, description="Cycle filter for execution"),
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Execute a report configuration with the specified cycle."""
    
    try:
        from app.services.report_execution import ReportExecutionService
        
        execution_service = ReportExecutionService(db, dw_db)
        result = execution_service.execute_report(report_id, cycle_filter)
        
        return result
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error executing report: {str(e)}")


@router.post("/reports/{report_id}/export/csv")
async def export_report_csv(
    report_id: int,
    cycle_filter: Optional[int] = Query(None, description="Cycle filter for execution"),
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Export report results as CSV."""
    
    try:
        from app.services.report_execution import ReportExecutionService
        from fastapi.responses import StreamingResponse
        import csv
        import io
        
        execution_service = ReportExecutionService(db, dw_db)
        result = execution_service.execute_report(report_id, cycle_filter)
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        headers = [col.header for col in result.columns]
        writer.writerow(headers)
        
        # Write data rows
        for row in result.rows:
            csv_row = []
            for col in result.columns:
                value = row.get(col.field, '')
                # Format value for CSV
                if isinstance(value, (int, float)) and value != value:  # NaN check
                    value = ''
                csv_row.append(str(value) if value is not None else '')
            writer.writerow(csv_row)
        
        # Create response
        output.seek(0)
        response = StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={result.report_name.replace(' ', '_')}.csv"
            }
        )
        
        return response
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting CSV: {str(e)}")


@router.post("/reports/{report_id}/export/excel")
async def export_report_excel(
    report_id: int,
    cycle_filter: Optional[int] = Query(None, description="Cycle filter for execution"),
    db: Session = Depends(get_db_session),
    dw_db: Session = Depends(get_dw_session)
):
    """Export report results as Excel."""
    
    try:
        from app.services.report_execution import ReportExecutionService
        from fastapi.responses import StreamingResponse
        import io
        
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="Excel export requires openpyxl. Install with: pip install openpyxl"
            )
        
        execution_service = ReportExecutionService(db, dw_db)
        result = execution_service.execute_report(report_id, cycle_filter)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = result.report_name[:31]  # Excel sheet name limit
        
        # Write headers with formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col_idx, col in enumerate(result.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data rows
        for row_idx, row in enumerate(result.rows, 2):
            for col_idx, col in enumerate(result.columns, 1):
                value = row.get(col.field)
                
                # Format value for Excel
                if value is not None:
                    if col.type == "number" and isinstance(value, (int, float)):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    elif col.type == "percentage" and isinstance(value, (int, float)):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value/100)
                        cell.number_format = '0.00%'
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=str(value))
                else:
                    ws.cell(row=row_idx, column=col_idx, value='')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={result.report_name.replace(' ', '_')}.xlsx"
            }
        )
        
        return response
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")


# === HELPER FUNCTIONS ===

def _get_available_fields_by_scope(db: Session) -> Dict[str, List[AvailableField]]:
    """Get available fields organized by scope."""
    
    # Raw field definitions
    raw_fields = [
        # Deal-level fields
        AvailableField(
            field_name="dl_nbr",
            display_name="Deal Number",
            description="Unique deal identifier",
            field_type="number",
            field_source=FieldSource.RAW_FIELD,
            category="Basic Info",
            is_default=True
        ),
        AvailableField(
            field_name="issr_cde",
            display_name="Issuer Code",
            description="Issuer identification code",
            field_type="text",
            field_source=FieldSource.RAW_FIELD,
            category="Basic Info",
            is_default=True
        ),
        AvailableField(
            field_name="cdi_file_nme",
            display_name="CDI File Name",
            description="CDI file name",
            field_type="text",
            field_source=FieldSource.RAW_FIELD,
            category="Basic Info",
            is_default=False
        ),
        # Tranche-level fields
        AvailableField(
            field_name="tr_id",
            display_name="Tranche ID",
            description="Tranche identifier",
            field_type="text",
            field_source=FieldSource.RAW_FIELD,
            category="Basic Info",
            is_default=True
        ),
        AvailableField(
            field_name="tr_cusip_id",
            display_name="Tranche CUSIP",
            description="Tranche CUSIP identifier",
            field_type="text",
            field_source=FieldSource.RAW_FIELD,
            category="Basic Info",
            is_default=False
        ),
        # Financial fields (both levels)
        AvailableField(
            field_name="tr_end_bal_amt",
            display_name="Ending Balance",
            description="Tranche ending balance amount",
            field_type="number",
            field_source=FieldSource.RAW_FIELD,
            category="Financial Data",
            is_default=True
        ),
        AvailableField(
            field_name="tr_prin_dstrb_amt",
            display_name="Principal Distribution",
            description="Principal distribution amount",
            field_type="number",
            field_source=FieldSource.RAW_FIELD,
            category="Financial Data",
            is_default=True
        ),
        AvailableField(
            field_name="tr_int_dstrb_amt",
            display_name="Interest Distribution",
            description="Interest distribution amount",
            field_type="number",
            field_source=FieldSource.RAW_FIELD,
            category="Financial Data",
            is_default=False
        ),
        AvailableField(
            field_name="tr_pass_thru_rte",
            display_name="Pass Through Rate",
            description="Interest pass-through rate",
            field_type="percentage",
            field_source=FieldSource.RAW_FIELD,
            category="Financial Data",
            is_default=False
        ),
    ]
    
    # Get saved calculations
    calc_repo = CalculationRepository(db)
    saved_calcs = calc_repo.get_all_calculations()
    
    calculation_fields = []
    for calc in saved_calcs:
        calculation_fields.append(
            AvailableField(
                field_name=calc.name,
                display_name=calc.name,
                description=calc.description or f"{calc.calculation_type} calculation",
                field_type="number",  # Most calculations return numbers
                field_source=FieldSource.SAVED_CALCULATION,
                category="Calculations",
                is_default=False,
                calculation_id=calc.id
            )
        )
    
    # Organize by scope
    deal_fields = [f for f in raw_fields if f.field_name not in ["tr_id", "tr_cusip_id"]]
    tranche_fields = raw_fields.copy()
    
    # Add calculations based on their aggregation level
    for calc_field in calculation_fields:
        calc = next((c for c in saved_calcs if c.id == calc_field.calculation_id), None)
        if calc:
            if calc.aggregation_level == "deal":
                deal_fields.append(calc_field)
            elif calc.aggregation_level == "tranche":
                tranche_fields.append(calc_field)
    
    return {
        "DEAL": deal_fields,
        "TRANCHE": tranche_fields
    }


def _get_operator_description(operator: FilterOperator) -> str:
    """Get description for filter operator."""
    descriptions = {
        FilterOperator.EQUALS: "Exactly matches the value",
        FilterOperator.NOT_EQUALS: "Does not match the value",
        FilterOperator.GREATER_THAN: "Greater than the value",
        FilterOperator.LESS_THAN: "Less than the value",
        FilterOperator.GREATER_THAN_OR_EQUAL: "Greater than or equal to the value",
        FilterOperator.LESS_THAN_OR_EQUAL: "Less than or equal to the value",
        FilterOperator.IN: "Matches any of the specified values",
        FilterOperator.NOT_IN: "Does not match any of the specified values",
        FilterOperator.CONTAINS: "Contains the text value",
        FilterOperator.NOT_CONTAINS: "Does not contain the text value",
        FilterOperator.IS_NULL: "Field has no value",
        FilterOperator.IS_NOT_NULL: "Field has a value"
    }
    return descriptions.get(operator, "")


def _convert_report_to_response(report: Report) -> ReportResponse:
    """Convert database report to response model."""
    from app.models.report_api_models import (
        ReportDealResponse, ReportTrancheResponse, 
        ReportFieldResponse, FilterConditionResponse
    )
    
    # Convert deals
    deals = []
    for deal in report.selected_deals:
        tranches = [
            ReportTrancheResponse(
                id=tranche.id,
                dl_nbr=tranche.dl_nbr,
                tr_id=tranche.tr_id
            )
            for tranche in deal.selected_tranches
        ]
        
        deals.append(
            ReportDealResponse(
                id=deal.id,
                dl_nbr=deal.dl_nbr,
                selected_tranches=tranches
            )
        )
    
    # Convert fields
    fields = [
        ReportFieldResponse(
            id=field.id,
            field_name=field.field_name,
            display_name=field.display_name,
            field_type=field.field_type,
            field_source=FieldSource(field.field_source),
            calculation_id=field.calculation_id,
            is_required=field.is_required
        )
        for field in report.selected_fields
    ]
    
    # Convert filter conditions
    conditions = [
        FilterConditionResponse(
            id=condition.id,
            field_name=condition.field_name,
            operator=FilterOperator(condition.operator),
            value=condition.value
        )
        for condition in report.filter_conditions
    ]
    
    return ReportResponse(
        id=report.id,
        name=report.name,
        description=report.description,
        scope=ReportScope(report.scope),
        created_by=report.created_by,
        created_date=report.created_date,
        updated_date=report.updated_date,
        is_active=report.is_active,
        selected_deals=deals,
        selected_fields=fields,
        filter_conditions=conditions
    )


def _convert_report_to_summary(report: Report) -> ReportSummaryResponse:
    """Convert database report to summary response."""
    return ReportSummaryResponse(
        id=report.id,
        name=report.name,
        description=report.description,
        scope=ReportScope(report.scope),
        created_by=report.created_by,
        created_date=report.created_date.isoformat(),
        updated_date=report.updated_date.isoformat(),
        deal_count=len(report.selected_deals),
        tranche_count=sum(len(deal.selected_tranches) for deal in report.selected_deals),
        field_count=len(report.selected_fields)
    )